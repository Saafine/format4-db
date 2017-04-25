#!/usr/bin/env python
# -*- coding: utf-8 -*-
# enable UTF-8 characters in the console

import fdb
import translate
from openpyxl import load_workbook
import datetime
from PyQt4 import QtCore, QtGui
from configparser import ConfigParser


class Current:
    fdb = None
    sheet = 'Plyty'
    filled_row = 3  # row number of the first row with data
    row_data = []
    unique_id_group = ''
    order_id = ''
    order_num = ''
    updated = False


class Config:
    parser = ConfigParser()
    file = 'settings.ini'
    parser.read(file)

    def write(self):
        with open(self.file, 'w') as configfile:
            self.parser.write(configfile)

    def update(self, section, key, value):
        self.parser.set(section, key, value)
        self.write()

    def remove(self, section):
        self.parser.remove_section(section)
        self.parser.add_section(section)
        self.write()


def unique_id():
    now = datetime.datetime.now()
    micro = str(now.microsecond)[:3]
    Current.unique_id_group = str(now.day) + str(now.month) + str(now.year) + str(now.hour) + str(now.minute) + micro  # unique id cannot exceed 20 characters due to database table column limiations


class Database:
    def __init__(self, path='baza/Baza.fdb', user='SYSDBA', password='masterkey'):  # standard firebird password
        self.con = fdb.connect(database=path, user=user, password=password)
        self.cur = self.con.cursor()
        self.table_format = 'FORMATKI'
        self.table_group = 'ZESTAWY'
        self.table_format_group = 'ZESTAWY_FORMATKI'
        self.table_boards = 'PLYTY'
        self.table_boards_columns = 'ID_PLYTY, OPIS'
        self.sheet = 'Plyty'
        self.format_columns = translate.get_table_headers(self.table_format, self.con, self.cur)  # check for available table headers, this fixes version issues
        self.group_columns = translate.get_table_headers(self.table_group, self.con, self.cur)
        self.format_group_columns = translate.get_table_headers(self.table_format_group, self.con, self.cur)
        self.boards_values = translate.get_columns_values(self.table_boards, self.table_boards_columns, self.con, self.cur)
        try:
            Config().remove('db_boards')
            for i in range(len(self.boards_values)):
                Config().update('db_boards', self.boards_values[i][1], self.boards_values[i][0])
        except:
            print('Updating settings.ini for existing boards failed')


class Settings:
    try:
        # read values from a section
        database_path = Config.parser.get('path', 'database')
        Current.fdb = Database(database_path)
    except:
        database_path = 'Nie znaleziono bazy danych.'


class OptionsDialog(QtGui.QDialog):
    def __init__(self, parent=None):
        super(OptionsDialog, self).__init__(parent)
        self.setGeometry(50, 50, 1000, 500)  # window dimensions
        self.setWindowTitle('Format Export - Opcje')  # window title
        self.setWindowIcon(QtGui.QIcon('img/icon.png'))  # window icon
        self.setWindowFlags(self.windowFlags() ^ QtCore.Qt.WindowContextHelpButtonHint)  # remove question mark button


class Window(QtGui.QMainWindow):
    def __init__(self):
        super(Window, self).__init__()  # or super().__init__() in 3.X python
        self.setGeometry(50, 50, 300, 300)  # window position + dimensions
        self.setWindowTitle('Format Export')  # window title
        self.setWindowIcon(QtGui.QIcon('img/icon.png'))  # window icon

        # Fonts
        font1 = QtGui.QFont()
        font1.setPointSize(10)
        font1.setBold(True)
        font1.setWeight(75)

        # Labels
        order_nr_tit = QtGui.QLabel('Numer zamówienia:')
        order_nr_tit.setFont(font1)
        self.order_nr_val = QtGui.QLabel('brak')
        self.order_nr_val.setFont(font1)
        path_format_tit = QtGui.QLabel('Ścieżka do formatek')
        self.path_format_val = QtGui.QLabel('wczytaj formatki')
        path_db_tit = QtGui.QLabel('Ścieżka bazy')
        self.path_db_val = QtGui.QLabel(Settings.database_path)
        # Buttons
        btn_open_format = QtGui.QPushButton('Otwórz formatki')
        btn_update_db = QtGui.QPushButton('Aktualizuj Formatyzację 5')
        # Checkboxes
        self.chk_add_as_group = QtGui.QCheckBox('Dodaj, jako zestaw')
        self.chk_add_as_group.setChecked(True)
        # Layout
        wid = QtGui.QWidget(self)
        self.setCentralWidget(wid)
        layout = QtGui.QGridLayout()
        layout.addWidget(order_nr_tit, 0, 0)
        layout.addWidget(self.order_nr_val, 0, 1)
        layout.addWidget(path_format_tit, 1, 0)
        layout.addWidget(self.path_format_val, 1, 1)
        layout.addWidget(path_db_tit, 2, 0)
        layout.addWidget(self.path_db_val, 2, 1)
        layout.addWidget(btn_open_format, 3, 0)
        layout.addWidget(btn_update_db, 3, 1)
        layout.addWidget(self.chk_add_as_group, 4, 0)
        layout.setSpacing(5)  # padding
        wid.setLayout(layout)

        mainMenu = self.menuBar()
        self.statusBar()
        # File menu file starts
        fileMenu = mainMenu.addMenu('&Plik')
        # Sub menu - open database
        action_db = QtGui.QAction('&Otwórz bazę danych', self)
        action_db.setShortcut('Ctrl+D')  # run action with buttons clicked
        action_db.setStatusTip('Otwórz bazę danych Firebird, np. C:\Formatyzacja 5\Bazy\Baza.fdb')  # status bar
        # Sub menu - open format
        action_format = QtGui.QAction('&Otwórz formatki', self)
        action_format.setShortcut('Ctrl+F')
        action_format.setStatusTip('Otwórz plik Excel z formatkami, w formacie .xls, .xlsx')
        # Sub menu - close app
        action_close = QtGui.QAction('&Zakończ', self)
        action_close.setShortcut('Ctrl+Q')
        action_close.setStatusTip('Zamknij aplikację')
        # Add file sub menus to menu
        fileMenu.addAction(action_format)
        fileMenu.addAction(action_db)
        fileMenu.addSeparator()  # adds line seperator in menu
        fileMenu.addAction(action_close)
        # File menu file ends

        # Sub menu - open help
        action_help = QtGui.QAction('&Pomoc', self)
        action_help.setStatusTip('Otwórz poomoc')  # status bar

        # Signals
        # --menu
        action_db.triggered.connect(self.database_open)
        action_format.triggered.connect(self.format_open)
        action_close.triggered.connect(self.close_application)
        action_help.triggered.connect(self.help_open)

        # --btn
        btn_open_format.clicked.connect(self.format_open)
        btn_update_db.clicked.connect(self.btn_database)

        # Child windows
        self.OptionsWindow = OptionsDialog(self)

    def options_open(self):
        self.OptionsWindow.exec_()

    def help_open(self):
        self.OptionsWindow.exec_()

    def showMessageBox(self, title, msg):
        QtGui.QMessageBox.information(self, title, msg)

    def database_open(self):
        path = QtGui.QFileDialog.getOpenFileName(self, 'Open File', '', 'Baza danych (*.fdb *.db)')
        if path == '':
            return

        Current.updated = False  # if db reselected, allow adding formats
        try:
            Config().update('path', 'database', str(path))
        except:
            print('Settings.ini not found')
        try:
            Current.fdb = Database(path)
            self.path_db_val.setText(path)
        except Exception:
                self.showMessageBox('Błąd', 'Nie udało się otworzyć bazy danych')
                print('Nie udało się otworzyć bazy danych')

    def format_open(self):
        unique_id()  # set unique id for opened format list
        path = QtGui.QFileDialog.getOpenFileName(self, 'Open File', '', 'Excel Woorkbook (*.xlsx *.xls *.xlsm)')
        if path == '':
            return
        Current.updated = False
        try:
            download_format(path)
            self.path_format_val.setText(str(path))
            self.order_nr_val.setText(str(Current.order_id))
        except:
            print('Nie udało się otworzyć formatek')

    def btn_database(self):
        add_as_group = self.chk_add_as_group.isChecked()
        if Current.updated:
            self.showMessageBox('Błąd', 'Formatki istnieją w bazie.')
            return
        if Current.fdb is None:
            self.showMessageBox('Błąd', 'Wczytaj bazę danych.')
            return
        elif Current.row_data == []:
            self.showMessageBox('Błąd', 'Wczytaj listę formatek.')
            return
        try:
            update_database(add_as_group)
            Current.updated = True
            self.showMessageBox('Formatyzacja', 'Baza danych została zaktualizowana.')
        except:
            self.showMessageBox('Błąd', 'Nie udało się zaaktualizować bazy danych.')

    def close_application(self):
        choice = QtGui.QMessageBox.question(self, 'Zakończ', 'Czy chcesz zamknąć program?', QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)

        if choice == QtGui.QMessageBox.Yes:
            sys.exit()
        else:
            pass


class RowImported:
    def __init__(self, element, quantity, length, width, material, surface, symbol):  # run every time when creating object
        # import from file values
        self.element = translate.x_utf(element)
        self.quantity = translate.x_utf(quantity)
        self.length = translate.x_utf(length)
        self.width = translate.x_utf(width)
        self.material = translate.x_utf(material)
        self.surface = translate.x_utf(surface)
        self.symbol = translate.x_utf(symbol)
        # export to database values
        self.desc_out = str(self.symbol)+'-'   # 'opis': 'BN1-000'
        self.thickness_out = str(translate.extract_number(self.material))  # 'grubosc': '28'
        if self.thickness_out == '' or self.thickness_out is None:
            self.thickness_out = '1'
        self.symbol_out = str(self.symbol) + ' - ' + str(self.element) + ' '  # 'symbol': 'BN1 - polka 000'
        self.board_id_out = str(translate.board_id(self.material))


def download_format(path):
    #  open excel sheet
    wb = load_workbook(filename=path, read_only=True)
    sheet = wb[Current.sheet]

    # check how many rows are filled, read data and create row objects
    Current.order_id = sheet.cell(row=1, column=1).value  # get order id from 1st column, 1st cell
    if Current.order_id == '' or Current.order_id is None:
        Current.order_id == 'XY 000'
    Current.order_num = translate.extract_number(Current.order_id)
    if Current.order_num == '' or Current.order_num is None:
        Current.order_num == '000'

    try:
        while sheet.cell(row=Current.filled_row, column=1).value is not None:
            temp_element = sheet.cell(row=Current.filled_row, column=1).value
            temp_quantity = sheet.cell(row=Current.filled_row, column=2).value
            temp_length = sheet.cell(row=Current.filled_row, column=3).value
            temp_width = sheet.cell(row=Current.filled_row, column=4).value
            temp_material = sheet.cell(row=Current.filled_row, column=5).value
            temp_surface = sheet.cell(row=Current.filled_row, column=6).value
            temp_symbol = sheet.cell(row=Current.filled_row, column=7).value
            Current.row_data += [RowImported(temp_element, temp_quantity, temp_length, temp_width, temp_material, temp_surface, temp_symbol)]
            Current.filled_row += 1
    except:  # fix for some versions of openpyxl that cant determine empty row
        Current.filled_row = 3
        max = sheet.max_row
        while Current.filled_row <= max:
            temp_element = sheet.cell(row=Current.filled_row, column=1).value
            temp_quantity = sheet.cell(row=Current.filled_row, column=2).value
            temp_length = sheet.cell(row=Current.filled_row, column=3).value
            temp_width = sheet.cell(row=Current.filled_row, column=4).value
            temp_material = sheet.cell(row=Current.filled_row, column=5).value
            temp_surface = sheet.cell(row=Current.filled_row, column=6).value
            temp_symbol = sheet.cell(row=Current.filled_row, column=7).value
            Current.row_data += [RowImported(temp_element, temp_quantity, temp_length, temp_width, temp_material, temp_surface, temp_symbol)]
            Current.filled_row += 1


def update_database(add_as_group):
    if add_as_group:
        group = translate.table_variables().data_group(Current.fdb.group_columns)  # get standard database values for existing column names
        group['ID_ZESTAWU'] = str(Current.unique_id_group)
        group['OPIS'] = str(Current.order_id)
        group['SYMBOL'] = str(Current.order_id)
        query_gr = translate.query_syntax(group)  # translate dictionary into query
        insert_group = 'insert into ' + Current.fdb.table_group + ' (' + query_gr[0] + ') values (' + query_gr[1] + ')'  # full query string
        Current.fdb.cur.execute(insert_group)  # execute query
        data_grouped = translate.table_variables().data_format_group(Current.fdb.format_group_columns)

    data = translate.table_variables().data_format(Current.fdb.format_columns)   # get standard database values for individual boards
    for r in range(len(Current.row_data)):
        # set individual values for each board
        data['ID_FORMATKI'] = str(Current.unique_id_group) + str(r)
        data['OPIS'] = str(Current.row_data[r].desc_out)+str(Current.order_num)
        data['DLUGOSC'] = str(Current.row_data[r].length)
        data['SZEROKOSC'] = str(Current.row_data[r].width)
        data['GRUBOSC'] = str(Current.row_data[r].thickness_out)
        data['SYMBOL'] = str(Current.row_data[r].symbol_out)+str(Current.order_num)
        data['CECHA'] = str(Current.order_id)
        data['GRUPA'] = str(Current.order_id)
        data['ID_PLYTY'] = str(Current.row_data[r].board_id_out)
        query_format = translate.query_syntax(data)  # translate dictionary into query
        insert_format = 'insert into ' + Current.fdb.table_format + ' (' + query_format[0] + ') values (' + query_format[1] + ')'  # full query string
        Current.fdb.cur.execute(insert_format)  # execute query
        if add_as_group:
            data_grouped['ID_FORMATKI'] = data['ID_FORMATKI']
            data_grouped['STAN_DEF'] = Current.row_data[r].quantity
            data_grouped['ID_Z_FORMATKI'] = str(Current.unique_id_group) + 'G' + str(r)
            data_grouped['ID_ZESTAWU'] = str(Current.unique_id_group)
            query_data_gr = translate.query_syntax(data_grouped)
            insert_format_grouped = 'insert into ' + Current.fdb.table_format_group + ' (' + query_data_gr[0] + ') values (' + query_data_gr[1] + ')'
            Current.fdb.cur.execute(insert_format_grouped)  # execute query
    Current.fdb.con.commit()  # flush


if __name__ == "__main__":
    import sys
    app = QtGui.QApplication(sys.argv)
    MainWindow = Window()
    MainWindow.show()
    sys.exit(app.exec_())



